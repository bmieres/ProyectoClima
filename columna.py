from openpyxl import load_workbook
# The source xlsx file is named as source.xlsx
wb=load_workbook("source.xlsx")

ws = wb.active
first_column = ws['A']

# Print the contents
for x in xrange(len(first_column)): 
    print(first_column[x].value) 